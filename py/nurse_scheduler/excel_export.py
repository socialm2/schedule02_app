# -*- coding: utf-8 -*-
"""엑셀 출력 (설계서 §7, G10~G12)."""
from __future__ import annotations

from typing import List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .calendar_utils import DayInfo, DAY_WEEKDAY, DAY_SATURDAY, holiday_count
from .models import MonthSchedule, Shift, NIGHT_SHIFTS, WORK_SHIFTS

SHIFT_FILLS = {
    Shift.A8: "D9D9D9",
    Shift.A9: "BFBFBF",
    Shift.D: "BDD7EE",
    Shift.E: "F8CBAD",
    Shift.N: "1F3864",
    Shift.NK: "7030A0",
    Shift.PRN: "C6E0B4",
    Shift.AL: "FFE699",
    Shift.EDU: "B4C7E7",
    Shift.BEREAVE: "808080",
    Shift.CELEBRATE: "F4B6C2",
    Shift.AL_1H: "FFF2CC",
    Shift.AL_2H: "FFF2CC",
    Shift.AL_3H: "FFF2CC",
    Shift.AL_HALF: "FFF2CC",
    Shift.OFFICIAL: "D0CECE",
    Shift.SICK: "F8C9C4",
    Shift.LEAVE: "BFBFBF",
    Shift.PROMO_EDU: "DDEBF7",
    Shift.SLEEP_OFF: "9DC3E6",
    Shift.TW: "A9D18E",
    Shift.MIL: "C9C9C9",
}
WHITE_FONT_SHIFTS = {Shift.N, Shift.NK, Shift.BEREAVE}
WEEKEND_FILL = PatternFill("solid", fgColor="FCE4EC")
SUB_FILL = PatternFill("solid", fgColor="FFD966")
THIN = Border(*[Side(style="thin", color="BFBFBF")] * 4)
CENTER = Alignment(horizontal="center", vertical="center")


def _count_eq(rng: str, code: str) -> str:
    """원티드 접미사(*)를 무시하고 code와 정확히 일치하는 칸 수를 세는 수식 조각.

    COUNTIF의 "*" 와일드카드는 접두사가 겹치는 코드(N vs NK 등)를 이중 계산하므로
    대신 SUBSTITUTE로 접미사를 지운 뒤 정확히 비교한다."""
    return f'SUMPRODUCT(--(SUBSTITUTE({rng},"*","")="{code}"))'


def export_excel(sch: MonthSchedule, days: List[DayInfo], path: str):
    """근무표 출력 (병원 표준 양식 호환).

    열 구성: 이름 | Lv | 잔휴(전월) | 잔휴(이번달) | 1일...말일 |
             D | E | N | ® | ⓡ | 금월 | T연 | R연 | 부서
    ®/ⓡ/T연/R연/부서는 우리 엔진이 추적하지 않는 병원 HR 항목이라 빈칸으로 둔다
    (표준 서식과의 열 자리는 맞추되 값은 지어내지 않음). 금월 = holiday_count()
    (이 달의 휴일수), 잔휴 = 전월 잔휴 + (금월 휴일수 - 이번달 순수오프일수).
    파트장 행 다음에 'A' 팀 구분 행을 넣어 병원 양식의 팀 구획 표기를 반영한다
    (Team B는 우리 엔진이 관리하는 개념이 아니라 생략)."""
    wb = Workbook()
    ws = wb.active
    ws.title = f"{sch.year}-{sch.month:02d}"
    off_bal_col1, off_bal_col2 = 3, 4  # 잔휴(전월)/잔휴(이번달)
    first_day_col = 5  # E열부터 1일
    last_day_col = first_day_col + sch.num_days - 1
    hol = holiday_count(days)
    sum_cols = ["D", "E", "N", "®", "ⓡ", "금월", "T연", "R연", "부서"]
    _write_schedule_sheet(ws, sch, days, hol, first_day_col, last_day_col,
                          off_bal_col1, off_bal_col2, sum_cols)
    wb.save(path)


def _write_schedule_sheet(ws, sch: MonthSchedule, days: List[DayInfo], hol: int,
                          first_day_col: int, last_day_col: int,
                          off_bal_col1: int, off_bal_col2: int, sum_cols: List[str]):
    nd = sch.num_days
    ws.cell(1, 1, f"{sch.year}년 {sch.month}월 근무표").font = Font(bold=True, size=14)
    ws.cell(2, 1, "이름").font = Font(bold=True)
    ws.cell(2, 2, "Lv").font = Font(bold=True)
    ws.cell(2, off_bal_col1, "잔휴").font = Font(bold=True)
    ws.cell(3, off_bal_col1, "전월").font = Font(bold=True, size=9)
    ws.cell(2, off_bal_col2, "잔휴").font = Font(bold=True)
    ws.cell(3, off_bal_col2, "이후").font = Font(bold=True, size=9)
    for c in (off_bal_col1, off_bal_col2):
        ws.cell(2, c).alignment = CENTER
        ws.cell(3, c).alignment = CENTER
        ws.column_dimensions[get_column_letter(c)].width = 6
    for d in range(nd):
        c = first_day_col + d
        di = days[d]
        cell = ws.cell(2, c, d + 1)
        dow = ws.cell(3, c, di.dow_name)
        for x in (cell, dow):
            x.alignment = CENTER
            x.font = Font(bold=True)
            if di.is_substitute:
                x.fill = SUB_FILL
            elif di.day_type != DAY_WEEKDAY:
                x.fill = WEEKEND_FILL
        ws.column_dimensions[get_column_letter(c)].width = 4.5
    for i, name in enumerate(sum_cols):
        c = last_day_col + 1 + i
        ws.cell(2, c, name).font = Font(bold=True)
        ws.cell(2, c).alignment = CENTER
        ws.column_dimensions[get_column_letter(c)].width = 6

    ordered = sorted(sch.staff, key=lambda s: (not s.is_partjang,))
    partjang_rows = [s for s in ordered if s.is_partjang]
    others = [s for s in ordered if not s.is_partjang]

    def write_staff_row(row, s):
        ws.cell(row, 1, s.id)
        ws.cell(row, 2, s.level).alignment = CENTER
        off_before = sch.carryover[s.id].off_balance
        ws.cell(row, off_bal_col1, round(off_before, 2)).alignment = CENTER
        for d in range(nd):
            v = sch.grid[s.id][d]
            if v == Shift.OFF:
                text = "X" if (s.id, d) in sch.wanted else "/"
            else:
                text = str(v) if v else ""
                if v and (s.id, d) in sch.wanted:
                    text += "*"
            c = ws.cell(row, first_day_col + d, text)
            c.alignment = CENTER
            c.border = THIN
            if v in SHIFT_FILLS:
                c.fill = PatternFill("solid", fgColor=SHIFT_FILLS[v])
                if v in WHITE_FONT_SHIFTS:
                    c.font = Font(color="FFFFFF")
        a = get_column_letter(first_day_col)
        b = get_column_letter(last_day_col)
        rng = f"{a}{row}:{b}{row}"
        off_after = round(off_before + hol - sch.offs_in_month(s.id), 2)
        ws.cell(row, off_bal_col2, off_after).alignment = CENTER
        ws.cell(row, last_day_col + 1, f'={_count_eq(rng, "D")}').alignment = CENTER
        ws.cell(row, last_day_col + 2, f'={_count_eq(rng, "E")}').alignment = CENTER
        ws.cell(row, last_day_col + 3,
                f'={_count_eq(rng, "N")}+{_count_eq(rng, "NK")}').alignment = CENTER
        ws.cell(row, last_day_col + 6, hol).alignment = CENTER  # 금월

    row = 4
    for s in partjang_rows:
        write_staff_row(row, s)
        row += 1
    # 'A' 팀 구분 행 (병원 양식 호환용 — Team B는 우리 엔진이 관리하지 않아 생략)
    ws.cell(row, 1, "A").font = Font(bold=True, color="808080")
    row += 1
    nurse_top = row
    for s in others:
        write_staff_row(row, s)
        row += 1
    nurse_bot = row - 1

    labels = [("D", ["D"]),
              ("E", ["E"]),
              ("N", ["N", "NK"]),
              ("prn", ["prn", "8A", "9A"])]
    sum_start = row + 1
    for i, (name, codes) in enumerate(labels):
        rr = sum_start + i
        ws.cell(rr, 1, f"{name} 계").font = Font(bold=True)
        for d in range(nd):
            col = get_column_letter(first_day_col + d)
            rng = f"{col}{nurse_top}:{col}{nurse_bot}"
            formula = "=" + "+".join(_count_eq(rng, code) for code in codes)
            cell = ws.cell(rr, first_day_col + d, formula)
            cell.alignment = CENTER
            cell.border = THIN

    rr = sum_start + len(labels)
    ws.cell(rr, 1, "Lv 평균").font = Font(bold=True)
    for d in range(nd):
        levels = [s.level for s in sch.staff
                  if not s.is_partjang and sch.grid[s.id][d] in WORK_SHIFTS]
        val = round(sum(levels) / len(levels), 2) if levels else ""
        cell = ws.cell(rr, first_day_col + d, val)
        cell.alignment = CENTER
        cell.border = THIN

    # B팀(신입) — 이름만 표시, 나머지 칸은 전부 빈칸(파트장이 직접 배정)
    if sch.team_b_names:
        rr += 2
        ws.cell(rr, 1, "B").font = Font(bold=True, color="808080")
        for name in sch.team_b_names:
            rr += 1
            ws.cell(rr, 1, name)

    ws.freeze_panes = ws.cell(4, first_day_col)
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 4


# ---------------------------------------------------------------- OCS 원본 형식 출력

def _cell_shift_text(sch: MonthSchedule, sid: str, d: int):
    """OFF는 X(원티드)/－ (일반) 표시로 바꾸고, 그 외는 원티드 접미사(*)를 붙인다.
    반환: (표시 텍스트, 실제 Shift 값 — 색칠용)."""
    v = sch.grid[sid][d]
    if v == Shift.OFF:
        text = "X" if (sid, d) in sch.wanted else "/"
    else:
        text = str(v) if v else ""
        if v and (sid, d) in sch.wanted:
            text += "*"
    return text, v


def export_excel_ocs(sch: MonthSchedule, days: List[DayInfo], path: str):
    """근무표 출력 — 병원 OCS가 실제로 내보내는 원본과 같은 모양(간호스케줄).

    열 구성: 순번 | 성명 | (빈칸, 원본의 익명화 코드는 재현 불가) | 잔휴(전월) |
             잔휴(이후) | 1일...말일 | D | E | N | ® | ⓡ | 금월 | T연 | R연 | 부서 | Lv
    T연/R연/부서/®/ⓡ는 우리 엔진이 추적하지 않는 HR 데이터라 빈칸. Lv는 원본에
    없는 우리 쪽 추가 정보라 맨 뒤에 붙인다. 원본처럼 하단 합계행은 없다(원본
    파일 자체가 그렇게 끝남) — 요약이 필요하면 기본 export_excel()을 쓴다.
    load_prev_month_schedule_xlsx()의 병원 OCS 폴백 파서로 그대로 재업로드 가능."""
    wb = Workbook()
    ws = wb.active
    ws.title = f"{sch.year}-{sch.month:02d}"
    nd = sch.num_days
    off_bal_col1, off_bal_col2 = 4, 5
    first_day_col = 6
    last_day_col = first_day_col + nd - 1
    lv_col = last_day_col + 10  # D,E,N,®,ⓡ,금월,T연,R연,부서(9) 다음
    hol = holiday_count(days)

    ws.cell(1, 2, "간호스케줄").font = Font(bold=True, size=14)
    ws.cell(2, 2, f"조회 : {sch.year}년{sch.month:02d}월").font = Font(bold=True)
    ws.cell(3, off_bal_col1, "전월").font = Font(bold=True, size=9)
    ws.cell(4, 2, "성  명").font = Font(bold=True)
    ws.cell(4, off_bal_col1, "잔휴").font = Font(bold=True, size=9)
    ws.cell(4, off_bal_col2, "잔휴").font = Font(bold=True, size=9)
    for c in (off_bal_col1, off_bal_col2):
        ws.cell(3, c).alignment = CENTER
        ws.cell(4, c).alignment = CENTER
        ws.column_dimensions[get_column_letter(c)].width = 6
    for d in range(nd):
        c = first_day_col + d
        di = days[d]
        cell = ws.cell(3, c, d + 1)
        dow = ws.cell(4, c, di.dow_name)
        for x in (cell, dow):
            x.alignment = CENTER
            x.font = Font(bold=True)
            if di.is_substitute:
                x.fill = SUB_FILL
            elif di.day_type != DAY_WEEKDAY:
                x.fill = WEEKEND_FILL
        ws.column_dimensions[get_column_letter(c)].width = 4.5
    for i, name in enumerate(["D", "E", "N", "®", "ⓡ", "금월", "T연", "R연", "부서"]):
        c = last_day_col + 1 + i
        ws.cell(3, c, name).font = Font(bold=True)
        ws.cell(3, c).alignment = CENTER
        ws.column_dimensions[get_column_letter(c)].width = 6
    ws.cell(3, lv_col, "Lv").font = Font(bold=True)
    ws.cell(3, lv_col).alignment = CENTER
    ws.column_dimensions[get_column_letter(lv_col)].width = 5

    def write_row(row, s, seq):
        if seq is not None:
            ws.cell(row, 1, seq).alignment = CENTER
        ws.cell(row, 2, s.id)
        off_before = sch.carryover[s.id].off_balance
        ws.cell(row, off_bal_col1, round(off_before, 2)).alignment = CENTER
        for d in range(nd):
            text, v = _cell_shift_text(sch, s.id, d)
            c = ws.cell(row, first_day_col + d, text)
            c.alignment = CENTER
            c.border = THIN
            if v in SHIFT_FILLS:
                c.fill = PatternFill("solid", fgColor=SHIFT_FILLS[v])
                if v in WHITE_FONT_SHIFTS:
                    c.font = Font(color="FFFFFF")
        a, b = get_column_letter(first_day_col), get_column_letter(last_day_col)
        rng = f"{a}{row}:{b}{row}"
        off_after = round(off_before + hol - sch.offs_in_month(s.id), 2)
        ws.cell(row, off_bal_col2, off_after).alignment = CENTER
        ws.cell(row, last_day_col + 1, f'={_count_eq(rng, "D")}').alignment = CENTER
        ws.cell(row, last_day_col + 2, f'={_count_eq(rng, "E")}').alignment = CENTER
        ws.cell(row, last_day_col + 3,
                f'={_count_eq(rng, "N")}+{_count_eq(rng, "NK")}').alignment = CENTER
        ws.cell(row, last_day_col + 6, hol).alignment = CENTER  # 금월
        ws.cell(row, lv_col, s.level).alignment = CENTER

    ordered = sorted(sch.staff, key=lambda s: (not s.is_partjang,))
    partjang_rows = [s for s in ordered if s.is_partjang]
    others = [s for s in ordered if not s.is_partjang]

    row = 5
    for s in partjang_rows:
        write_row(row, s, None)  # 파트장은 원본에서도 순번이 없음
        row += 1
    ws.cell(row, 2, "A").font = Font(bold=True, color="808080")
    row += 1
    seq = 0
    for s in others:
        seq += 1
        write_row(row, s, seq)
        row += 1
    if sch.team_b_names:
        ws.cell(row, 2, "B").font = Font(bold=True, color="808080")
        row += 1
        for name in sch.team_b_names:
            seq += 1
            ws.cell(row, 1, seq).alignment = CENTER
            ws.cell(row, 2, name)
            row += 1

    ws.freeze_panes = ws.cell(5, first_day_col)
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 8
    wb.save(path)


# ---------------------------------------------------------------- 연간근무표 (다월 자동 연동용, 구 명칭 "인원표")

STAT_FILL = PatternFill("solid", fgColor="E2EFDA")
NK_ROW_FILL = PatternFill("solid", fgColor="F2E9F7")
GROUP_FONT = Font(bold=True, size=11, color="808080")
BOLD = Font(bold=True)


def export_staff_table_xlsx(staff, stats: dict, annual_dates: list,
                             annual_grid: dict, path: str,
                             last_reflected: str = ""):
    """'연간근무표' 내보내기 — 로스터 + 누적통계(형평성 참고용) + 연간 근무 그리드(1/1~, 참고용).

    staff: List[Staff] (현재 인원 명단, 이 순서/구성대로 출력)
    stats: {staff_id: {"night","workday","off","weekend_night",
                        "blocks_2","blocks_3","months","recent_night_score"}}
    annual_dates: [datetime.date, ...] — 1/1부터 이어진 연간 그리드 날짜(오름차순)
    annual_grid: {staff_id: [code, ...]} — annual_dates와 나란히
    """
    from .excel_input import STAFF_TABLE_STATIC_COLS, STAFF_TABLE_STAT_KEYS, STAFF_TABLE_STAT_LABELS
    from .calendar_utils import WEEKDAY_NAMES

    wb = Workbook()
    ws = wb.active
    ws.title = "연간근무표"

    n_static = len(STAFF_TABLE_STATIC_COLS)
    n_stat = len(STAFF_TABLE_STAT_KEYS)
    first_day_col = 1 + n_static + n_stat  # 1-based

    title = "연간근무표"
    if last_reflected:
        title += f" — {last_reflected}까지 반영"
    ws.cell(1, 1, title).font = Font(bold=True, size=13)

    ws.cell(2, 1, "[ 인원 정보 ]").font = GROUP_FONT
    ws.cell(2, 1 + n_static, "[ 누적 통계 (형평성 참고용) ]").font = GROUP_FONT
    if annual_dates:
        ws.cell(2, first_day_col, "[ 연간 근무표 (참고용, 1/1 리셋) ]").font = GROUP_FONT

    header_row = 3
    for i, name in enumerate(STAFF_TABLE_STATIC_COLS):
        c = ws.cell(header_row, 1 + i, name)
        c.font = BOLD
        c.alignment = CENTER
    for i, name in enumerate(STAFF_TABLE_STAT_LABELS):
        c = ws.cell(header_row, 1 + n_static + i, name)
        c.font = BOLD
        c.alignment = CENTER
        c.fill = STAT_FILL

    dow_row = header_row + 1
    for i, d in enumerate(annual_dates):
        c = first_day_col + i
        cell = ws.cell(header_row, c, d)
        cell.number_format = "m/d"
        cell.font = Font(bold=True, size=9)
        cell.alignment = CENTER
        dow = ws.cell(dow_row, c, WEEKDAY_NAMES[d.weekday()])
        dow.font = Font(bold=True, size=9)
        dow.alignment = CENTER
        if d.weekday() >= 5:
            cell.fill = WEEKEND_FILL
            dow.fill = WEEKEND_FILL
        ws.column_dimensions[get_column_letter(c)].width = 4.2

    row0 = dow_row + 1
    for r, s in enumerate(staff):
        row = row0 + r
        is_nk = "night_only" in (s.flags or [])
        note = []
        if is_nk:
            note.append("야간전담")
        if "pregnant" in (s.flags or []):
            note.append("임부(야간 금지)")
        elif "no_night" in (s.flags or []):
            note.append("야간금지")
        static_vals = [r + 1, s.id, s.role, f"Lv{s.level}",
                       ",".join(s.allowed_shifts), ", ".join(note)]
        for i, v in enumerate(static_vals):
            c = ws.cell(row, 1 + i, v)
            c.alignment = CENTER
            if is_nk:
                c.fill = NK_ROW_FILL

        st = stats.get(s.id, {})
        for i, key in enumerate(STAFF_TABLE_STAT_KEYS):
            v = st.get(key, 0)
            c = ws.cell(row, 1 + n_static + i, v)
            c.alignment = CENTER
            c.fill = NK_ROW_FILL if is_nk else STAT_FILL

        codes = annual_grid.get(s.id, [])
        for i in range(len(annual_dates)):
            code = codes[i] if i < len(codes) else ""
            c = ws.cell(row, first_day_col + i, code)
            c.alignment = CENTER
            c.font = Font(size=9)
            if code in SHIFT_FILLS:
                c.fill = PatternFill("solid", fgColor=SHIFT_FILLS[code])
                if code in WHITE_FONT_SHIFTS:
                    c.font = Font(size=9, color="FFFFFF")

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 9
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 22
    if annual_dates:
        ws.freeze_panes = ws.cell(row0, first_day_col)
    wb.save(path)

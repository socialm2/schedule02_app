# -*- coding: utf-8 -*-
"""엑셀 입력 형식 (원본 설계서 0-2/0-3 표 양식).

시트 구성:
  설정     — 항목/값 (연도, 월, 월최대야간, 월신청상한, 공휴일, 대체공휴일)
  인원     — 순번 | 이름 | 직급 | 숙련도 | 가능근무 | 비고
  최소인력 — 근무 | 평일 | 토요일 | 일요일공휴일  (D/E/N/prn 행, 8A 행은 무시)
  신청     — 이름 | 날짜 | 근무유형 | 우선순위   (선택, 행 순서 = 선착순)
  전월이월 — 이름 | 마지막근무 | 연속근무일수 | 이월OFF일수 | 말일연속야간일수 (선택)

비고 키워드: "야간전담" → night_only, "임부" → pregnant, "야간금지"/"야간 금지" → no_night

  연간근무표 — 다월 자동 연동용(선택 사용, 구 명칭 "인원표"). 순번|이름|직급|숙련도|가능근무|비고 +
           누적통계(최근야간점수|누적야간|누적오프|누적주말야간|누적2일블록|누적3일블록|반영개월수) +
           연간 근무 그리드(1/1~, 참고용). export_staff_table_xlsx()로 내보내고
           load_staff_table_xlsx()로 다시 읽어 다음 달 생성에 이어붙인다.
"""
from __future__ import annotations

from typing import Dict, List, Optional

from .calendar_utils import kr_holidays_for_month
from .models import Shift, parse_shift

STAFF_TABLE_STATIC_COLS = ["순번", "이름", "직급", "숙련도", "가능근무", "비고"]
STAFF_TABLE_STAT_KEYS = ["night", "workday", "off", "weekend_night",
                          "blocks_2", "blocks_3", "months", "recent_night_score",
                          "cum_d", "cum_e", "cum_prn", "cum_weekend_holiday_off"]
STAFF_TABLE_STAT_LABELS = ["누적야간", "누적근무일", "누적오프", "누적주말야간",
                            "누적2일블록", "누적3일블록", "반영개월수",
                            "최근야간점수(참고용)",
                            "누적D", "누적E", "누적prn(8A)", "누적주말휴일오프"]


class ExcelInputError(Exception):
    pass


# ---------------------------------------------------------------- 읽기

def _sheet_rows(ws) -> List[list]:
    return [[c for c in row] for row in ws.iter_rows(values_only=True)]


def _find_header(rows, first_col_name):
    for i, r in enumerate(rows):
        if r and str(r[0] or "").strip() == first_col_name:
            return i
    return None


def _parse_level(v) -> int:
    s = str(v).strip()
    if s.lower().startswith("lv"):
        s = s[2:]
    try:
        return int(float(s))
    except ValueError:
        raise ExcelInputError(f"숙련도 해석 불가: {v!r}")


def _parse_flags(note: str) -> List[str]:
    note = (note or "").replace(" ", "")
    flags = []
    if "야간전담" in note:
        flags.append("night_only")
    if "임부" in note:
        flags.append("pregnant")
    if "야간금지" in note and "pregnant" not in flags:
        flags.append("no_night")
    return flags


def _split_list(v) -> List[str]:
    if v is None:
        return []
    return [x.strip() for x in str(v).replace("·", ",").split(",") if x.strip()]


def load_input_xlsx(path: str) -> dict:
    """입력 엑셀 → 설정 dict (JSON 입력과 동일 스키마)."""
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True)

    for need in ("설정", "인원", "최소인력"):
        if need not in wb.sheetnames:
            raise ExcelInputError(f"필수 시트 누락: {need}")

    # -------- 설정
    kv: Dict[str, object] = {}
    for row in wb["설정"].iter_rows(values_only=True):
        if row and row[0] is not None and str(row[0]).strip() not in ("항목", ""):
            kv[str(row[0]).strip()] = row[1] if len(row) > 1 else None
    try:
        year = int(kv["연도"])
        month = int(kv["월"])
    except (KeyError, TypeError, ValueError):
        raise ExcelInputError("설정 시트에 연도/월이 없습니다")

    # -------- 인원
    rows = _sheet_rows(wb["인원"])
    h = _find_header(rows, "순번")
    if h is None:
        raise ExcelInputError("인원 시트에 '순번' 헤더가 없습니다")
    staff = []
    for r in rows[h + 1:]:
        if not r or r[1] is None or not str(r[1]).strip():
            continue
        name = str(r[1]).strip()
        role = str(r[2] or "간호사").strip()
        level = _parse_level(r[3])
        allowed = _split_list(r[4])
        note = str(r[5] or "") if len(r) > 5 else ""
        for a in allowed:
            parse_shift(a)  # 검증
        staff.append({
            "id": name, "role": role, "level": level,
            "allowed_shifts": allowed, "flags": _parse_flags(note),
        })
    if not staff:
        raise ExcelInputError("인원 시트가 비어 있습니다")

    # -------- 최소인력
    rows = _sheet_rows(wb["최소인력"])
    h = _find_header(rows, "근무")
    if h is None:
        raise ExcelInputError("최소인력 시트에 '근무' 헤더가 없습니다")
    mins = {"weekday": {}, "saturday": {}, "sunday_holiday": {}}
    for r in rows[h + 1:]:
        if not r or r[0] is None:
            continue
        key = str(r[0]).strip()
        if key not in ("D", "E", "N", "prn"):
            continue  # 8A 등은 무시 (파트장 자동 배정)
        mins["weekday"][key] = int(r[1] or 0)
        mins["saturday"][key] = int(r[2] or 0)
        mins["sunday_holiday"][key] = int(r[3] or 0)

    # -------- 신청 (선택)
    requests = []
    if "신청" in wb.sheetnames:
        rows = _sheet_rows(wb["신청"])
        h = _find_header(rows, "이름")
        if h is not None:
            for r in rows[h + 1:]:
                if not r or r[0] is None or not str(r[0]).strip():
                    continue
                date = r[1]
                if hasattr(date, "strftime"):
                    date = date.strftime("%Y-%m-%d")
                requests.append({
                    "staff_id": str(r[0]).strip(),
                    "date": str(date).strip()[:10],
                    "type": str(r[2]).strip(),
                    "priority": int(r[3] or 1) if len(r) > 3 else 1,
                })

    # -------- 전월이월 (선택)
    carry = {}
    if "전월이월" in wb.sheetnames:
        rows = _sheet_rows(wb["전월이월"])
        h = _find_header(rows, "이름")
        if h is not None:
            for r in rows[h + 1:]:
                if not r or r[0] is None or not str(r[0]).strip():
                    continue
                trailing = int(r[4] or 0) if len(r) > 4 else 0
                carry[str(r[0]).strip()] = {
                    "last_shift_type": str(r[1] or "OFF").strip(),
                    "consecutive_work_days": int(r[2] or 0),
                    "night_block_remaining_off": int(r[3] or 0),
                    "night_block_in_progress": trailing == 1,
                    "trailing_night_count": trailing,
                }

    # -------- B팀(신입, 선택) — 이름만. 스케줄링에 관여하지 않고 출력 시 이름만 표시.
    team_b_names = []
    if "B팀" in wb.sheetnames:
        rows = _sheet_rows(wb["B팀"])
        h = _find_header(rows, "이름")
        if h is not None:
            for r in rows[h + 1:]:
                if r and r[0] is not None and str(r[0]).strip():
                    team_b_names.append(str(r[0]).strip())

    nk_count = sum(1 for s in staff if "night_only" in s["flags"])
    holidays = _split_list(kv.get("공휴일"))
    substitute_holidays = _split_list(kv.get("대체공휴일"))
    if not holidays:
        # 사용자가 공휴일을 안 채웠으면(엑셀 업로드 경로는 화면의 "한국 달력 기준 자동 반영"
        # 캘린더 위젯을 안 거치므로) 알려진 한국 공휴일로 자동 채운다 — kr_holidays_for_month().
        holidays, auto_subs = kr_holidays_for_month(year, month)
        if not substitute_holidays:
            substitute_holidays = auto_subs
    return {
        "ward_id": str(kv.get("병동명", "") or ""),
        "year": year, "month": month,
        "staff": staff,
        "params": {
            "nk_count": int(kv.get("NK인원수", nk_count) or nk_count),
            "min_staff": mins,
            # 리더8A운용은 웹 UI에도 없는 설정이라 엑셀에서도 뺐다 — 항상 기본값(False)만 쓴다.
            "leader_8a_as_prn": False,
            "off_max_per_month": int(kv.get("월최대야간", 6) or 6),
            "max_requests_per_person": int(kv.get("월신청상한", 6) or 6),
            "holidays": holidays,
            "substitute_holidays": substitute_holidays,
            "advanced_track_staff": _split_list(kv.get("심화과정대상")),
            "team_b_names": team_b_names,
        },
        "prev_month_carryover": carry,
        "requests": requests,
    }


# ---------------------------------------------------------------- 월간근무표 형태 엑셀 공통 (제목행 + 이름|1일|2일|...)

def _parse_monthly_grid_shape(ws):
    """제목 행(A1) 'YYYY년 M월' + 2행 일자 헤더 + 4행부터 이름|근무... 모양 파싱.

    "월간근무표"(이 앱이 내보낸 근무표) 다운로드/전월이월 자동채움/원티드 업로드가
    전부 이 모양을 공유한다. 반환: (year, month, num_days, first_day_col).
    """
    import re
    title = str(ws.cell(1, 1).value or "")
    m = re.match(r"(\d{4})\s*년\s*(\d{1,2})\s*월", title)
    if not m:
        raise ExcelInputError(
            "제목 행(A1)에서 'YYYY년 M월' 형식을 찾지 못했습니다 — "
            "월간근무표와 같은 모양(이름|1일|2일|...)인지 확인하세요."
        )
    year, month = int(m.group(1)), int(m.group(2))
    # 일자 열: 예전 포맷(3열부터 1일)과 잔휴 2열이 추가된 신규 포맷(5열부터 1일) 둘 다
    # 읽을 수 있도록, 2행에서 "1"이 나오는 열을 찾아 첫 일자 열로 삼는다(고정 열번호
    # 대신 동적 탐지 — 선두 열이 몇 개든 견고하게 동작).
    first_day_col = None
    for col in range(1, 10):
        if ws.cell(2, col).value == 1:
            first_day_col = col
            break
    if first_day_col is None:
        raise ExcelInputError("2행에서 1일(숫자 1) 열을 찾지 못했습니다.")
    col = first_day_col
    while isinstance(ws.cell(2, col).value, (int, float)):
        col += 1
    num_days = col - first_day_col
    if not (28 <= num_days <= 31):
        raise ExcelInputError(f"일자 열을 인식하지 못했습니다 (인식된 일수: {num_days})")
    return year, month, num_days, first_day_col


def _try_parse_hospital_ocs_shape(ws):
    """병원 OCS 시스템이 직접 내보낸 근무표 원본(예: '간호스케줄') 모양 인식 시도.

    확인된 실제 구조: 1행 '간호스케줄', 2행 '조회 : YYYY년MM월', 3행에 일자 헤더
    ('전월' 라벨 + 1~말일 + D/E/N/®/ⓡ/금월/T연/R연/부서), 4행에 요일(같은 열에
    '성  명' 라벨), 5행부터 데이터(파트장 특별행 → 'A' 팀 구분 행 → 순번 매긴 일반
    직원 → 'B' 팀 구분 행 → B팀 직원). 순번은 1열, 성명은 2열.

    우리 자체 포맷(_parse_monthly_grid_shape)과 달리 실패해도 예외를 던지지 않고
    None을 돌려준다 — 호출 쪽에서 우리 포맷을 먼저 시도하고 실패했을 때만 이
    함수로 폴백하기 위함(서로 다른 두 실제 파일 모양을 각각 안전하게 인식).
    반환: {"year","month","num_days","first_day_col","header_row","name_col"} 또는 None.
    """
    import re
    year = month = None
    for r in range(1, 4):
        for c in range(1, 7):
            v = str(ws.cell(r, c).value or "")
            m = re.search(r"(\d{4})\s*년\s*(\d{1,2})\s*월", v)
            if m:
                year, month = int(m.group(1)), int(m.group(2))
                break
        if year:
            break
    if year is None:
        return None

    header_row = first_day_col = num_days = None
    for r in range(1, 7):
        for c in range(1, 10):
            if ws.cell(r, c).value == 1:
                c2, n = c, 1
                while ws.cell(r, c2 + 1).value == n + 1:
                    c2 += 1
                    n += 1
                if 28 <= n <= 31:
                    header_row, first_day_col, num_days = r, c, n
                    break
        if header_row:
            break
    if header_row is None:
        return None

    # 성명 열: 일자 헤더 다음 행(요일 행)에서 '성명'/'이름' 라벨이 있는 열을 찾는다.
    name_col = None
    for c in range(1, first_day_col):
        label = "".join(str(ws.cell(header_row + 1, c).value or "").split())
        if "성명" in label or "이름" in label:
            name_col = c
            break
    if name_col is None:
        return None

    return {"year": year, "month": month, "num_days": num_days,
            "first_day_col": first_day_col, "header_row": header_row,
            "name_col": name_col}


def _find_grid_sheet(wb):
    """월간근무표 모양의 시트를 찾는다.

    엑셀에서 마지막으로 보고 있던 시트(wb.active)를 우선 시도하되, 안내/설명용
    시트가 같이 들어있어 그게 활성 시트로 저장된 경우에도 문제없이 읽히도록
    나머지 시트를 전부 훑어 'YYYY년 M월' 형식의 진짜 데이터 시트를 찾는다.
    반환: (ws, year, month, num_days, first_day_col)"""
    candidates = [wb.active] + [ws for ws in wb.worksheets if ws is not wb.active]
    last_err = None
    for ws in candidates:
        try:
            year, month, num_days, first_day_col = _parse_monthly_grid_shape(ws)
            return ws, year, month, num_days, first_day_col
        except ExcelInputError as e:
            last_err = e
    raise last_err or ExcelInputError("월간근무표 모양의 시트를 찾지 못했습니다.")


def _is_team_divider_row(ws, row: int, name_col: int, first_day_col: int, num_days: int) -> bool:
    """'A'/'B' 팀 구분 행 판별(export_excel도, 병원 OCS 원본도 이름 칸에만 'A'/'B'를
    쓰고 나머지(일자 칸)는 비워둔다).

    이름이 정확히 'A'/'B' 한 글자이면서 일자 칸이 전부 비어 있을 때만 구분 행으로
    본다 — 직원 이름이 우연히 'A'/'B'인 경우를 오인하지 않도록 이중으로 확인."""
    name = str(ws.cell(row, name_col).value or "").strip()
    if name not in ("A", "B"):
        return False
    return all(ws.cell(row, first_day_col + d).value in (None, "") for d in range(num_days))


def _read_grid_rows(ws, name_col: int, first_day_col: int, num_days: int,
                    data_start_row: int, has_off_balance: bool):
    """이름 열/일자 시작 열이 주어졌을 때 공통 행 읽기 루프.

    'A'/'B' 팀 구분 행은 건너뛰고, has_off_balance면 잔휴(이후, 일자 시작 열
    바로 앞 칸)도 함께 읽는다. 반환: (grid, off_balance)."""
    grid: Dict[str, List[str]] = {}
    off_balance: Dict[str, float] = {}
    row = data_start_row
    while True:
        name = ws.cell(row, name_col).value
        if name is None or str(name).strip() == "":
            break
        if _is_team_divider_row(ws, row, name_col, first_day_col, num_days):
            row += 1
            continue
        sid = str(name).strip()
        shifts = []
        for d in range(num_days):
            v = ws.cell(row, first_day_col + d).value
            shifts.append(str(v).strip() if v not in (None, "") else "OFF")
        grid[sid] = shifts
        if has_off_balance:
            v = ws.cell(row, first_day_col - 1).value
            try:
                off_balance[sid] = float(v)
            except (TypeError, ValueError):
                pass
        row += 1
    return grid, off_balance


# ---------------------------------------------------------------- 지난달 확정 근무표 읽기 (이월/히스토리용)

def load_prev_month_schedule_xlsx(path: str) -> dict:
    """월간근무표(이 앱이 내보낸 근무표 엑셀, 또는 병원 OCS가 직접 내보낸 원본)를
    읽어 그리드로 복원.

    먼저 우리 자체 포맷(제목 'YYYY년 M월'로 시작)으로 시도하고, 어느 시트에서도
    맞지 않으면 병원 OCS 원본 모양('간호스케줄'/'조회 : ...')으로 재시도한다 —
    서로 다른 두 실제 파일 모양을 각각 안전하게 인식하고, 어느 쪽에도 맞지
    않으면(형식이 또 다르면) 조용히 잘못 읽는 대신 에러를 낸다.

    수기로 값만 고친 파일도 읽을 수 있도록, 서식/수식이 아니라 셀 값만 본다.
    'A'/'B' 팀 구분 행은 건너뛴다. 잔휴 열이 있으면 '잔휴(이후)' 값도 함께 읽어
    다음 달 이월용으로 돌려준다.
    반환: {"year", "month", "num_days", "grid": {staff_id: [shift_str, ...]},
           "off_balance": {staff_id: float}}  # 잔휴 열이 없으면 빈 dict
    """
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    try:
        ws, year, month, num_days, first_day_col = _find_grid_sheet(wb)
        grid, off_balance = _read_grid_rows(ws, 1, first_day_col, num_days, 4,
                                            has_off_balance=first_day_col >= 5)
    except ExcelInputError as our_err:
        grid = {}
        year = month = num_days = None
        for ws in wb.worksheets:
            shape = _try_parse_hospital_ocs_shape(ws)
            if shape is None:
                continue
            year, month, num_days = shape["year"], shape["month"], shape["num_days"]
            grid, off_balance = _read_grid_rows(
                ws, shape["name_col"], shape["first_day_col"], num_days,
                shape["header_row"] + 2, has_off_balance=True)
            if grid:
                break
        if not grid:
            raise our_err

    if not grid:
        raise ExcelInputError("근무표에서 인원 행을 찾지 못했습니다.")

    return {"year": year, "month": month, "num_days": num_days, "grid": grid,
            "off_balance": off_balance}


# ---------------------------------------------------------------- 원티드 신청 읽기 (월간근무표 모양 + 신청 표기)

_WANTED_WORK_CODES = {"D": "D", "E": "E", "N": "N", "8A": "8A", "9A": "9A",
                      "NK": "NK", "T": "T", "TW": "TW"}
_WANTED_AL_CODES = {"연": "연차", "연차": "연차",
                    "연1": "연1", "연2": "연2", "연3": "연3", "연4": "연4"}
# 세부 사유가 명확한 휴가류는 각자의 코드로, 그 외(포상휴가 등 미분류 특수코드)는
# 뭉뚱그려 OFF 희망으로 해석한다.
_WANTED_LEAVE_CODES = {
    "조": "조", "경": "경", "공": "공", "병": "병", "휴": "휴", "승": "승",
    "군": "군", "S/": "S/",
}
_WANTED_OFF_CODES = {"X", "OFF", "포"}


def load_wanted_grid_xlsx(path: str) -> dict:
    """월간근무표와 같은 모양(이름|1일|2일|...) 엑셀에서 원티드 신청만 읽어온다.

    표기 규칙: "D*"/"E*"/"N*"/"8A*"/"9A*"/"NK*"/"T*"/"TW*"(또는 별표 없이도 인식) = 그 근무 희망,
    "조"/"경"/"공"/"병"/"휴"/"승"/"군"/"S/"/"연"/"연차"/"연1"/"연2"/"연3"/"연4" = 각자의
    휴가유형 희망, "X"/그 외 미분류 표기(포상휴가 등)는 OFF 희망으로 해석한다. 빈 칸은 건너뛴다.

    반환: {"year", "month", "requests": [{"staff_id","date","type"}, ...],
           "unknown_marks": [str, ...]}  # 인식 못 한 표시(참고용)
    """
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    ws, year, month, num_days, first_day_col = _find_grid_sheet(wb)

    requests: List[dict] = []
    unknown = set()
    row = 4
    while True:
        name = ws.cell(row, 1).value
        if name is None or str(name).strip() == "":
            break
        if _is_team_divider_row(ws, row, 1, first_day_col, num_days):
            row += 1
            continue
        sid = str(name).strip()
        for d in range(num_days):
            v = ws.cell(row, first_day_col + d).value
            if v in (None, ""):
                continue
            v = str(v).strip()
            base = v[:-1] if v.endswith("*") else v
            if base in _WANTED_WORK_CODES:
                t = _WANTED_WORK_CODES[base]
            elif base in _WANTED_AL_CODES:
                t = _WANTED_AL_CODES[base]
            elif base in _WANTED_LEAVE_CODES:
                t = _WANTED_LEAVE_CODES[base]
            elif base in _WANTED_OFF_CODES:
                t = "OFF"
            else:
                unknown.add(v)
                continue
            requests.append({"staff_id": sid, "date": f"{year}-{month:02d}-{d + 1:02d}",
                             "type": t})
        row += 1

    if not requests and not unknown:
        raise ExcelInputError("표에서 원티드 표시를 하나도 찾지 못했습니다.")

    return {"year": year, "month": month, "requests": requests,
           "unknown_marks": sorted(unknown)}


# ---------------------------------------------------------------- 연간근무표 읽기 (다월 자동 연동용)

def load_staff_table_xlsx(path: str) -> dict:
    """'연간근무표' 파일 읽기 — 로스터 + 누적 통계(형평성 참고용) + 연간 근무 그리드(1/1~, 참고용).

    시트명은 "연간근무표"를 우선 찾고, 예전에 내보낸 파일(구 명칭 "인원표")도
    그대로 읽을 수 있도록 하위호환으로 받아준다.

    반환: {
      "staff": [{"id","role","level","allowed_shifts","flags"}, ...],
      "stats": {staff_id: {"recent_night_score","night","off","weekend_night",
                            "blocks_2","blocks_3","months"}},
      "annual_dates": ["YYYY-MM-DD", ...],   # 이미 쌓여있던 연간 그리드의 날짜(있으면)
      "annual_grid": {staff_id: [code, ...]},  # annual_dates와 나란히
    }
    """
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    sheet_name = "연간근무표" if "연간근무표" in wb.sheetnames else "인원표"
    if sheet_name not in wb.sheetnames:
        raise ExcelInputError("필수 시트 누락: 연간근무표")

    rows = _sheet_rows(wb[sheet_name])
    h = _find_header(rows, "순번")
    if h is None:
        raise ExcelInputError("연간근무표 시트에 '순번' 헤더가 없습니다")
    header_row = rows[h]
    n_static = len(STAFF_TABLE_STATIC_COLS)
    n_stat = len(STAFF_TABLE_STAT_KEYS)
    first_day_idx = n_static + n_stat  # 0-based

    annual_dates: List[str] = []
    i = first_day_idx
    while i < len(header_row) and hasattr(header_row[i], "strftime"):
        annual_dates.append(header_row[i].strftime("%Y-%m-%d"))
        i += 1

    def _num(r, i) -> float:
        v = r[i] if i < len(r) else None
        try:
            return float(v) if v not in (None, "") else 0.0
        except (TypeError, ValueError):
            return 0.0

    staff, stats, annual_grid = [], {}, {}
    for r in rows[h + 1:]:
        if not r or r[1] is None or not str(r[1]).strip():
            continue
        name = str(r[1]).strip()
        role = str(r[2] or "간호사").strip()
        level = _parse_level(r[3])
        allowed = _split_list(r[4])
        note = str(r[5] or "") if len(r) > 5 else ""
        for a in allowed:
            parse_shift(a)  # 검증
        staff.append({"id": name, "role": role, "level": level,
                       "allowed_shifts": allowed, "flags": _parse_flags(note)})

        stats[name] = {
            "night": int(_num(r, n_static + 0)),
            "workday": int(_num(r, n_static + 1)),
            "off": int(_num(r, n_static + 2)),
            "weekend_night": int(_num(r, n_static + 3)),
            "blocks_2": int(_num(r, n_static + 4)),
            "blocks_3": int(_num(r, n_static + 5)),
            "months": int(_num(r, n_static + 6)),
            "recent_night_score": round(_num(r, n_static + 7), 2),
            "cum_d": int(_num(r, n_static + 8)),
            "cum_e": int(_num(r, n_static + 9)),
            "cum_prn": int(_num(r, n_static + 10)),
            "cum_weekend_holiday_off": int(_num(r, n_static + 11)),
        }

        if annual_dates:
            codes = []
            for d in range(len(annual_dates)):
                idx = first_day_idx + d
                v = r[idx] if idx < len(r) else None
                codes.append(str(v).strip() if v not in (None, "") else "")
            annual_grid[name] = codes

    if not staff:
        raise ExcelInputError("연간근무표 시트가 비어 있습니다")

    return {"staff": staff, "stats": stats,
            "annual_dates": annual_dates, "annual_grid": annual_grid}


# ---------------------------------------------------------------- 쓰기

def write_input_xlsx(cfg: dict, path: str):
    """설정 dict → 입력 엑셀 (원본 0-2/0-3 표 양식)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    head = Font(bold=True)
    fill = PatternFill("solid", fgColor="DDEBE7")

    ws = wb.active
    ws.title = "설정"
    p = cfg.get("params", {})
    rows = [("항목", "값"),
            ("병동명", cfg.get("ward_id", "")),
            ("연도", cfg["year"]), ("월", cfg["month"]),
            ("월최대야간", p.get("off_max_per_month", 6)),
            ("월신청상한", p.get("max_requests_per_person", 6)),
            ("공휴일", ", ".join(p.get("holidays", []))),
            ("대체공휴일", ", ".join(p.get("substitute_holidays", [])))]
    for r in rows:
        ws.append(list(r))
    for c in ws[1]:
        c.font = head
        c.fill = fill
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 30

    ws = wb.create_sheet("인원")
    ws.append(["순번", "이름", "직급", "숙련도", "가능근무", "비고"])
    for c in ws[1]:
        c.font = head
        c.fill = fill
    for i, s in enumerate(cfg["staff"], 1):
        flags = s.get("flags", [])
        notes = []
        if "night_only" in flags:
            notes.append("야간전담")
        if "pregnant" in flags:
            notes.append("임부(야간 금지)")
        elif "no_night" in flags:
            notes.append("야간금지")
        if s["role"] == "파트장":
            notes.append("카운트 제외, 상근(평일)")
        ws.append([i, s["id"], s["role"], f"Lv{s['level']}",
                   ",".join(s["allowed_shifts"]), ", ".join(notes)])
    for col, w in (("A", 6), ("B", 12), ("C", 9), ("D", 8), ("E", 18), ("F", 24)):
        ws.column_dimensions[col].width = w

    ws = wb.create_sheet("최소인력")
    ws.append(["근무", "평일", "토요일", "일요일공휴일"])
    for c in ws[1]:
        c.font = head
        c.fill = fill
    m = p.get("min_staff", {})
    for key in ("D", "E", "N", "prn"):
        ws.append([key, m.get("weekday", {}).get(key, 0),
                   m.get("saturday", {}).get(key, 0),
                   m.get("sunday_holiday", {}).get(key, 0)])
    ws.append(["8A", "1(파트장)", 0, 0])

    reqs = cfg.get("requests") or []
    if reqs:
        ws = wb.create_sheet("신청")
        ws.append(["이름", "날짜", "근무유형", "우선순위"])
        for c in ws[1]:
            c.font = head
            c.fill = fill
        for r in reqs:
            ws.append([r["staff_id"], r["date"], r["type"], r.get("priority", 1)])
        ws.column_dimensions["B"].width = 12

    carry = cfg.get("prev_month_carryover") or {}
    if carry:
        ws = wb.create_sheet("전월이월")
        ws.append(["이름", "마지막근무", "연속근무일수", "이월OFF일수", "말일연속야간일수"])
        for c in ws[1]:
            c.font = head
            c.fill = fill
        for sid, v in carry.items():
            ws.append([sid, v.get("last_shift_type", "OFF"),
                       v.get("consecutive_work_days", 0),
                       v.get("night_block_remaining_off", 0),
                       v.get("trailing_night_count", 0)])
        for col in ("A", "B", "C", "D", "E"):
            ws.column_dimensions[col].width = 14

    team_b_names = p.get("team_b_names") or []
    if team_b_names:
        ws = wb.create_sheet("B팀")
        ws.append(["이름"])
        for c in ws[1]:
            c.font = head
            c.fill = fill
        for name in team_b_names:
            ws.append([name])
        ws.column_dimensions["A"].width = 14

    wb.save(path)

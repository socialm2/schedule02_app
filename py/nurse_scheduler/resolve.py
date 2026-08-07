# -*- coding: utf-8 -*-
"""재생성(re-solve) 프로세스.

리더가 수정한 근무표를 읽어 기계 생성본과 비교하고,
수정 칸을 고정한 채 나머지를 다시 배정한다.

흐름: 수정본 로드 → diff → 수정본 전체 하드/소프트 검사(피드백)
      → 사용자 확인 → 수정 칸 고정 재생성
"""
from __future__ import annotations

import json
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple

from .constraints import all_hard_checks, check_soft
from .generator import Generator, generate_best
from .models import Shift, Violation, parse_shift


@dataclass
class Edit:
    staff_id: str
    day: int          # 0-based
    old: Shift
    new: Shift

    def __str__(self):
        return f"{self.staff_id} {self.day + 1}일: {self.old} → {self.new}"


# ---------------------------------------------------------------- 로드

def load_schedule_json(path: str) -> Dict[str, List[str]]:
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    return data["schedule"] if "schedule" in data else data


def load_schedule_xlsx(path: str, staff_ids: List[str],
                       num_days: int) -> Dict[str, List[str]]:
    """엑셀 근무표 읽기: A열 이름, 일자 열부터 1일. 빈 칸은 OFF로 간주.

    일자 시작 열은 2행에서 값 1이 나오는 열을 찾아 동적으로 정한다 — 예전 포맷
    (3열부터 1일)과 잔휴 2열이 추가된 신규 포맷(5열부터 1일) 모두 지원."""
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    first_day_col = None
    for col in range(1, 10):
        if ws.cell(2, col).value == 1:
            first_day_col = col
            break
    if first_day_col is None:
        first_day_col = 3  # 못 찾으면 예전 포맷 기본값으로 후퇴
    known = set(staff_ids)
    grid: Dict[str, List[str]] = {}
    for row in ws.iter_rows(min_col=1, max_col=first_day_col - 1 + num_days):
        name = row[0].value
        if not isinstance(name, str) or name.strip() not in known:
            continue
        name = name.strip()
        vals = []
        for c in row[first_day_col - 1:first_day_col - 1 + num_days]:
            v = c.value
            v = str(v).strip() if v is not None else ""
            vals.append(v if v else "OFF")
        grid[name] = vals
    missing = known - set(grid)
    if missing:
        raise ValueError(f"엑셀에서 직원 행을 찾지 못함: {sorted(missing)}")
    return grid


def parse_grid(raw: Dict[str, List[str]], staff_ids: List[str],
               num_days: int) -> Dict[str, List[Shift]]:
    out = {}
    for sid in staff_ids:
        if sid not in raw:
            raise ValueError(f"근무표에 {sid} 행이 없음")
        row = raw[sid]
        if len(row) != num_days:
            raise ValueError(f"{sid}: 일수 {len(row)} ≠ {num_days}")
        out[sid] = [parse_shift(v) for v in row]
    return out


# ---------------------------------------------------------------- diff

def diff_schedules(base: Dict[str, List[Shift]],
                   edited: Dict[str, List[Shift]]) -> List[Edit]:
    edits = []
    for sid, row in edited.items():
        for d, v in enumerate(row):
            if base[sid][d] != v:
                edits.append(Edit(staff_id=sid, day=d,
                                  old=base[sid][d], new=v))
    return edits


# ---------------------------------------------------------------- 피드백

@dataclass
class Feedback:
    edits: List[Edit]
    hard: List[Violation]         # 수정본 전체의 하드 위반
    soft: List[Violation]
    hard_edit_related: List[Violation]
    soft_edit_related: List[Violation]

    @property
    def has_warnings(self) -> bool:
        return bool(self.hard_edit_related or self.soft_edit_related)


def _edit_related(v: Violation, edits: List[Edit]) -> bool:
    for e in edits:
        if v.staff_id and v.staff_id == e.staff_id:
            if v.day is None or abs(v.day - e.day) <= 2:
                return True
        elif not v.staff_id and v.day is not None and v.day == e.day:
            return True
    return False


def analyze_edits(cfg: dict, edited_grid: Dict[str, List[Shift]],
                  edits: List[Edit]) -> Feedback:
    """수정본을 그대로 검증해 수정 관련 이상을 골라낸다 (재생성 전 피드백)."""
    gen = Generator(cfg)
    sch = gen.sch
    for s in sch.staff:
        for d in range(sch.num_days):
            v = edited_grid[s.id][d]
            sch.grid[s.id][d] = v
            if v in (Shift.A8, Shift.A9) and s.is_leader:
                sch.leader_8a.add((s.id, d))
    hard = all_hard_checks(sch, gen.days, gen.params)
    soft = [v for v in check_soft(sch, gen.days, gen.params)
            if v.severity == "soft"]
    return Feedback(
        edits=edits, hard=hard, soft=soft,
        hard_edit_related=[v for v in hard if _edit_related(v, edits)],
        soft_edit_related=[v for v in soft if _edit_related(v, edits)],
    )


def format_feedback(fb: Feedback) -> str:
    L = ["===== 수정분 피드백 ====="]
    L.append(f"수정 칸: {len(fb.edits)}건")
    for e in fb.edits:
        L.append(f"  - {e}")
    L.append("")
    if fb.hard_edit_related:
        L.append(f"■ 수정 관련 하드 제약 이상: {len(fb.hard_edit_related)}건 ★ 확인 필요")
        for v in fb.hard_edit_related:
            L.append(f"  - [{v.rule}] {v.message}")
        L.append("  ※ 수정 칸은 고정한 채 재생성하므로, 재배정으로 해소 불가한 항목은"
                 " 최종 리포트에도 위반으로 남습니다.")
    else:
        L.append("■ 수정 관련 하드 제약 이상 없음")
    L.append("")
    if fb.soft_edit_related:
        L.append(f"■ 수정 관련 소프트 제약 이상: {len(fb.soft_edit_related)}건")
        for v in fb.soft_edit_related:
            L.append(f"  - [{v.rule}] {v.message}")
    else:
        L.append("■ 수정 관련 소프트 제약 이상 없음")
    other_hard = [v for v in fb.hard if v not in fb.hard_edit_related]
    if other_hard:
        L.append("")
        L.append(f"■ (참고) 수정과 무관한 하드 이상 {len(other_hard)}건 — 재생성 시 해소 시도")
    return "\n".join(L) + "\n"


# ---------------------------------------------------------------- 재생성

def resolve(cfg: dict, edits: List[Edit]):
    """수정 칸을 고정하고 나머지를 재생성. 하드 위반 잔존 시 재시도."""
    fixed: Dict[Tuple[str, int], Shift] = {
        (e.staff_id, e.day): e.new for e in edits
    }
    return generate_best(cfg, fixed_cells=fixed)
